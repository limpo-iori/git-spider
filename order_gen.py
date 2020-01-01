import openpyxl


def read_order(file_name):
    order_data = openpyxl.load_workbook(file_name)
    order_data_sheet = order_data['Sheet1']
    return order_data_sheet


def create_order():
    order_xlsx = openpyxl.Workbook()
    order_xlsx_header = order_xlsx.active
    order_xlsx_header.title = '订单主单'
    order_xlsx_body = order_xlsx.create_sheet('订单明细')

    order_xlsx_header.cell(1, 1, '订单号')
    order_xlsx_header.cell(1, 2, '身份证姓名')
    order_xlsx_header.cell(1, 3, '身份证号码')
    order_xlsx_header.cell(1, 4, '电话')
    order_xlsx_header.cell(1, 5, '省')
    order_xlsx_header.cell(1, 6, '市')
    order_xlsx_header.cell(1, 7, '区')
    order_xlsx_header.cell(1, 8, '地址')

    order_xlsx_body.cell(1, 1, '订单号')
    order_xlsx_body.cell(1, 2, '条码')
    order_xlsx_body.cell(1, 3, '货号')
    order_xlsx_body.cell(1, 4, '数量')

    return order_xlsx


if __name__ == '__main__':
    order_data_sheet = read_order('order.xlsx')
    sheet_books = int((order_data_sheet.max_row-1) / 100) + 1
    n = 0
    for sheet_book in range(sheet_books):
        order_xlsx = create_order()
        order_xlsx_header = order_xlsx['订单主单']
        order_xlsx_body = order_xlsx['订单明细']
        for row in range(2, 102):
            order_xlsx_header.cell(row, 1, value=order_data_sheet.cell(row + n, 1).value)
            order_xlsx_header.cell(row, 2, value=order_data_sheet.cell(row + n, 2).value)
            order_xlsx_header.cell(row, 3, value=order_data_sheet.cell(row + n, 3).value)
            order_xlsx_header.cell(row, 4, value=order_data_sheet.cell(row + n, 4).value)
            order_xlsx_header.cell(row, 5, value=order_data_sheet.cell(row + n, 5).value)
            order_xlsx_header.cell(row, 6, value=order_data_sheet.cell(row + n, 6).value)
            order_xlsx_header.cell(row, 7, value=order_data_sheet.cell(row + n, 7).value)
            order_xlsx_header.cell(row, 8, value=order_data_sheet.cell(row + n, 8).value)

            order_xlsx_body.cell(row, 1, value=order_data_sheet.cell(row + n, 1).value)
            order_xlsx_body.cell(row, 2, value=order_data_sheet.cell(row + n, 9).value)
            order_xlsx_body.cell(row, 3, value=order_data_sheet.cell(row + n, 10).value)
            order_xlsx_body.cell(row, 4, value=order_data_sheet.cell(row + n, 11).value)

        n = n + 100
        order_xlsx.save(str(n) + '.xlsx')
        print('数据生成成功'+str(n))
    print('数据生成完成')