import openpyxl


if __name__ == '__main__':
    order_data_sheet = openpyxl.load_workbook('order.xlsx')['Sheet1']
    sheet_books = int((order_data_sheet.max_row-1) / 100) + 1
    n = 0
    for sheet_book in range(sheet_books):
        order_xlsx = openpyxl.load_workbook('order_model.xlsx')
        order_xlsx_header = order_xlsx['订单主单']
        order_xlsx_body = order_xlsx['订单明细']
        for row in range(2, 102):
            order_xlsx_header.cell(row, 1, value=order_data_sheet.cell(row + n, 1).value)
            order_xlsx_header.cell(row, 2, value=order_data_sheet.cell(row + n, 2).value)
            order_xlsx_header.cell(row, 3, value=order_data_sheet.cell(row + n, 3).value)
            order_xlsx_header.cell(row, 4, value=order_data_sheet.cell(row + n, 4).value)
            order_xlsx_header.cell(row, 5, value=order_data_sheet.cell(row + n, 5).value)
            order_xlsx_header.cell(row, 6, value=order_data_sheet.cell(row + n, 6).value)
            order_xlsx_header.cell(row, 7, value=order_data_sheet.cell(row + n, 7).value)
            order_xlsx_header.cell(row, 8, value=order_data_sheet.cell(row + n, 8).value)

            order_xlsx_body.cell(row, 1, value=order_data_sheet.cell(row + n, 1).value)
            order_xlsx_body.cell(row, 2, value=order_data_sheet.cell(row + n, 9).value)
            order_xlsx_body.cell(row, 3, value=order_data_sheet.cell(row + n, 10).value)
            order_xlsx_body.cell(row, 4, value=order_data_sheet.cell(row + n, 11).value)

        n = n + 100
        order_xlsx.save(str(n) + '.xlsx')

        print('数据生成成功'+str(n))
    print('数据生成完成')