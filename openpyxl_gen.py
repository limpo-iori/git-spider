import openpyxl
model_data = openpyxl.load_workbook('model.xlsx')
model_sheet_data = model_data['Sheet1']
data_xlsx = openpyxl.Workbook()
data_sheet = data_xlsx.active
cursor_row = 0
for row in model_sheet_data.rows:
    package_name = row[0].value
    package_count = row[1].value
    package_goods = row[2].value
    for package in range(1,package_count+1):
        data_sheet.cell(row=package+cursor_row,column=1,value=package_name)
        data_sheet.cell(row=package+cursor_row,column=2,value=package_goods)
    cursor_row += package
    print('已生成数据包：'+str(package_name))
print('数据生成完毕，合计：'+str(cursor_row))
data_xlsx.save('data.xlsx')